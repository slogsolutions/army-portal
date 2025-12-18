# registration/admin.py
import csv
import json
from datetime import timedelta
import zipfile
import os as _os  # for urandom

from django import forms
from django.conf import settings
from django.contrib import admin
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.urls import reverse, path
from django.utils import timezone
from django.utils.html import format_html
import openpyxl
from openpyxl.utils import get_column_letter

from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives.ciphers.aead import AESGCM

from .models import CandidateProfile
from results.models import CandidateAnswer
from questions.models import QuestionPaper

from django.apps import apps
from django.db import transaction
from django.contrib import messages

# Use a custom admin index template to show dashboard buttons.
admin.site.index_template = "registration/admin_index.html"


def wipe_exam_data_view(request):
    """
    Admin-only view to delete all exam-related data while keeping:
      - accounts.User
      - registration.CandidateProfile
      - reference.Trade
    Only records are deleted; tables stay intact.
    """
    if not request.user.is_superuser:
        return HttpResponseForbidden("Not allowed.")

    if request.method == "POST":
        protected_models = {
            ("accounts", "User"),
            ("registration", "CandidateProfile"),
            ("reference", "Trade"),
        }
        allowed_apps = {
            "accounts",
            "registration",
            "questions",
            "centers",
            "exams",
            "results",
            "syncops",
            "reference",
        }

        with transaction.atomic():
            # First clear models that PROTECT-link to Question so Questions can be removed cleanly.
            try:
                from results.models import CandidateAnswer
                from exams.models import Answer as ExamAnswer

                CandidateAnswer.objects.all().delete()
                ExamAnswer.objects.all().delete()
            except Exception:
                # If these models are missing for any reason, don't block the wipe.
                pass

            # Now iterate over all models and delete everything except the protected ones.
            for model in apps.get_models():
                app_label = model._meta.app_label
                model_name = model.__name__

                if app_label not in allowed_apps:
                    continue
                if (app_label, model_name) in protected_models:
                    continue

                model.objects.all().delete()

        messages.success(
            request,
            "All data has been deleted except Users, Candidate Profiles, and Trades.",
        )
        from django.shortcuts import redirect

        return redirect("admin:index")

    # GET: show confirmation page
    from django.shortcuts import render

    return render(request, "registration/wipe_data_confirm.html")


# -------------------------
# Custom Admin Form with Validation
# -------------------------
class CandidateProfileAdminForm(forms.ModelForm):
    class Meta:
        model = CandidateProfile
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()

        # Update instance with form data before validation
        instance = self.instance
        for field_name in [
            "trade",
            "primary_practical_marks",
            "primary_viva_marks",
            # "secondary_practical_marks",
            # "secondary_viva_marks",
        ]:
            if field_name in cleaned_data:
                setattr(instance, field_name, cleaned_data[field_name])

        # Run model validation
        try:
            instance.full_clean()
        except ValidationError as e:
            if hasattr(e, "message_dict"):
                for field, messages in e.message_dict.items():
                    if isinstance(messages, list):
                        for message in messages:
                            self.add_error(field, message)
                    else:
                        self.add_error(field, messages)
            else:
                raise forms.ValidationError(str(e))

        return cleaned_data


# -------------------------
# CSV exporter (candidate answers)
# -------------------------
def export_candidate_answers(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="selected_candidates_answers.csv"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Army Number",
            "Candidate Name",
            "Paper Title",
            "Question ID",
            "Question Text",
            "Answer",
            "Category",
            "Submitted At",
        ]
    )

    answers = (
        CandidateAnswer.objects.filter(candidate__in=queryset)
        .select_related("candidate", "paper", "question")
        .order_by("candidate__id", "paper_id", "question_id")
    )

    for ans in answers:
        writer.writerow(
            [
                getattr(ans.candidate, "army_no", ""),
                ans.candidate.name if ans.candidate else "",
                getattr(ans.paper, "title", ""),
                getattr(ans.question, "id", ""),
                getattr(ans.question, "text", ""),
                getattr(ans, "answer", ""),
                ans.effective_category,  # computed: primary/secondary
                getattr(ans, "submitted_at", ""),
            ]
        )
    return response


export_candidate_answers.short_description = "Export selected candidates' answers to CSV"


# -------------------------
# Excel exporter (candidates)
# -------------------------
def export_candidates_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates"

    columns = [
        "Army No",
        "Rank",
        "Name",
        "Photo",
        "Trade",
        "DOB",
        "Father Name",
        "Date of Enrolment",
        "Aadhar Number",
        "Training Center",
        "District",
        "State",
        "Primary Qualification",
        "Primary Duration",
        "Primary Credits",
        # "Secondary Qualification",
        # "Secondary Duration",
        # "Secondary Credits",
        "NSQF Level",
        "Exam Center",
        "Shift",
        "Created At",
    ]

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num).value = column_title

    for row_num, candidate in enumerate(queryset, 2):
        # Safe access for optional fields
        photo = getattr(candidate, "photograph", None)
        photo_url = getattr(photo, "url", "") if photo else ""
        father_name = getattr(candidate, "father_name", "")
        aadhar_number = getattr(candidate, "aadhar_number", "")

        data = [
            candidate.army_no,
            candidate.rank,
            candidate.name,
            photo_url,
            getattr(candidate.trade, "name", str(candidate.trade)) if getattr(candidate, "trade", None) else "",
            candidate.dob,
            father_name,
            candidate.doe.strftime("%Y-%m-%d") if candidate.doe else "",
            aadhar_number,
            getattr(candidate, "training_center", ""),
            getattr(candidate, "district", ""),
            getattr(candidate, "state", ""),
            candidate.primary_qualification,
            candidate.primary_duration,
            candidate.primary_credits,
            # candidate.secondary_qualification,
            # candidate.secondary_duration,
            # candidate.secondary_credits,
            candidate.nsqf_level,
            candidate.exam_center,
            str(candidate.shift) if candidate.shift else "",
            candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else "",
        ]
        for col_num, cell_value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="candidates.xlsx"'
    wb.save(response)
    return response


export_candidates_excel.short_description = "Export selected candidates to Excel"


# -------------------------
# Helper: Build a multi-sheet workbook for .dat payload
# -------------------------
def _build_export_workbook(queryset):
    from openpyxl import Workbook
    from io import BytesIO

    from questions.models import QuestionPaper, ExamSession
    from results.models import CandidateAnswer

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # ✅ FINAL CLEAN HEADERS (NO EMPTY FIELDS)
    headers = [
        "S.No",
        "Name",
        "Center",
        "DOB",
        "Rank",
        "Category",
        "Trade_Type",
        "Trade",
        "Army_No",
        "Exam_Type",
        "Part",
        "Question",
        "Answer",
        "Correct_Answer",
        "Max_Marks",
    ]

    ws.append(headers)
    serial = 1

    for candidate in queryset:

        # Fetch exam sessions
        sessions = (
            ExamSession.objects
            .filter(user=candidate.user)
            .select_related("paper")
            .prefetch_related("examquestion_set__question")
            .order_by("-started_at")
        )

        # Fallback: if no session exists, use answered papers
        if not sessions.exists():
            papers = QuestionPaper.objects.filter(
                candidate_answers__candidate=candidate
            ).distinct()

            for paper in papers:
                questions = paper.questions.all().order_by("id")
                exam_type = "Secondary" if getattr(paper, "is_common", False) else "Primary"

                for q in questions:
                    ans = CandidateAnswer.objects.filter(
                        candidate=candidate,
                        paper=paper,
                        question=q
                    ).first()

                    row = [
                        serial,
                        candidate.name,
                        candidate.exam_center,
                        candidate.dob,
                        candidate.rank,
                        candidate.cat,               # ✅ Category
                        candidate.trade_type,        # ✅ Trade Type
                        candidate.trade.name if candidate.trade else "",
                        candidate.army_no,
                        exam_type,
                        q.part,
                        q.text,
                        ans.answer if ans and ans.answer is not None else "N/A",
                        getattr(q, "correct_answer", None),
                        q.marks if hasattr(q, "marks") else None,
                    ]

                    ws.append(row)
                    serial += 1

            continue

        # Normal flow: iterate sessions & assigned questions
        for session in sessions:
            paper = session.paper
            exam_type = (
                "Secondary"
                if paper and getattr(paper, "is_common", False)
                else "Primary"
            )

            for eq in session.questions:
                q = eq.question

                if paper:
                    ans = CandidateAnswer.objects.filter(
                        candidate=candidate,
                        paper=paper,
                        question=q
                    ).first()
                else:
                    ans = CandidateAnswer.objects.filter(
                        candidate=candidate,
                        question=q
                    ).first()

                row = [
                    serial,
                    candidate.name,
                    candidate.exam_center,
                    candidate.dob,
                    candidate.rank,
                    candidate.cat,               # ✅ Category
                    candidate.trade_type,        # ✅ Trade Type
                    candidate.trade.name if candidate.trade else "",
                    candidate.army_no,
                    exam_type,
                    q.part,
                    q.text,
                    ans.answer if ans and ans.answer is not None else "N/A",
                    getattr(q, "correct_answer", None),
                    q.marks if hasattr(q, "marks") else None,
                ]

                ws.append(row)
                serial += 1

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()



# -------------------------
# Crypto helper: encrypt bytes → .dat (salt + iv + ciphertext)
# -------------------------
def _encrypt_bytes_to_dat(data: bytes, passphrase: str) -> bytes:
    if not passphrase:
        raise ValueError("Missing CONVERTER_PASSPHRASE in settings.")

    salt = _os.urandom(16)
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=salt,
        iterations=100000,
    )
    key = kdf.derive(passphrase.encode("utf-8"))

    iv = _os.urandom(12)
    aesgcm = AESGCM(key)
    ciphertext = aesgcm.encrypt(iv, data, None)  # AAD=None

    # Layout: salt (16) || iv (12) || ciphertext (includes auth tag)
    return salt + iv + ciphertext


# -------------------------
# DAT exporter (encrypted .xlsx inside, converter-compatible)
# -------------------------
def export_candidates_dat(modeladmin, request, queryset):
    xlsx_bytes = _build_export_workbook(queryset)

    passphrase = getattr(settings, "CONVERTER_PASSPHRASE", None)
    if not passphrase:
        return HttpResponseBadRequest(
            "Server missing CONVERTER_PASSPHRASE; set it in settings or env."
        )

    dat_bytes = _encrypt_bytes_to_dat(xlsx_bytes, passphrase)

    from centers.models import Center

    center = Center.objects.first()

    if center:
        safe_exam_center = "".join(c if c.isalnum() else "_" for c in center.exam_Center)
        safe_comd = "".join(c if c.isalnum() else "_" for c in center.comd)
        filename = f"{safe_comd}_{safe_exam_center}.dat"
    else:
        ts = timezone.now().strftime("%Y%m%d%H%M%S")
        filename = f"candidates_export_{ts}.dat"

    response = HttpResponse(dat_bytes, content_type="application/octet-stream")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# Changed label: this will be displayed as the action/button text
export_candidates_dat.short_description = "Export All Exam Data"


# -------------------------
# Export candidate images as ZIP
# -------------------------
def export_candidate_images(modeladmin, request, queryset):
    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for candidate in queryset:
            photo = getattr(candidate, "photograph", None)
            if photo:
                try:
                    file_path = photo.path
                    ext = file_path[file_path.rfind(".") :] if "." in file_path else ""
                    filename = f"{candidate.army_no}_{candidate.name}{ext}"
                    zip_file.write(file_path, arcname=filename)
                except Exception:
                    continue

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="candidate_images.zip"'
    return response


# changed label as requested
export_candidate_images.short_description = "Export All Photos"


def export_all_candidate_images(modeladmin, request):
    from io import BytesIO

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for candidate in CandidateProfile.objects.all():
            photo = getattr(candidate, "photograph", None)
            if photo:
                try:
                    file_path = photo.path
                    ext = file_path[file_path.rfind(".") :] if "." in file_path else ""
                    filename = f"{candidate.army_no}_{candidate.name}{ext}"
                    zip_file.write(file_path, arcname=filename)
                except Exception:
                    continue

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="all_candidate_images.zip"'
    return response


# -------------------------
# NEW: Export marks (primary/secondary viva & practical) to Excel for selected queryset
# -------------------------
def export_marks_excel(modeladmin, request, queryset):
    """
    Export a simple Excel sheet with marks columns for the selected candidates.
    Columns: Army No, Name, Trade, Primary Viva, Primary Practical, Training Center, Exam Center, Created At
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"

    columns = [
        "Army No",
        "Name",
        "Trade",
        "Primary Viva Marks",
        "Primary Practical Marks",
        # "Secondary Viva Marks",
        # "Secondary Practical Marks",
        "Training Center",
        "Exam Center",
        "Created At",
    ]

    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num).value = column_title

    for row_num, candidate in enumerate(queryset, 2):
        trade_obj = getattr(candidate, "trade", None)
        trade_name = getattr(trade_obj, "name", str(trade_obj)) if trade_obj else ""
        row = [
            candidate.army_no,
            candidate.name,
            trade_name,
            candidate.primary_viva_marks,
            candidate.primary_practical_marks,
            # candidate.secondary_viva_marks,
            # candidate.secondary_practical_marks,
            getattr(candidate, "training_center", ""),
            candidate.exam_center,
            candidate.created_at.strftime("%Y-%m-%d %H:%M") if candidate.created_at else "",
        ]
        for col_num, cell_value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="candidate_marks.xlsx"'
    wb.save(response)
    return response


export_marks_excel.short_description = "Export Viva-Prac Marks"


# -------------------------
# Admin Registration
# -------------------------
@admin.register(CandidateProfile)
class CandidateProfileAdmin(admin.ModelAdmin):
    form = CandidateProfileAdminForm

    # default (non-PO) list
    list_display = ("army_no", "name", "user", "rank", "trade", "shift", "created_at")
    # base declaration; we will set this per-request in changelist_view
    list_editable = ()
    list_filter = ("trade", "training_center")

    # all actions declared; we'll filter them per user in get_actions
    actions = [
        export_candidate_answers,
        export_candidates_excel,
        export_candidates_dat,
        export_candidate_images,
        export_marks_excel,  # include new marks export as an action
    ]

    # ---------- helpers ----------
    def _is_po(self, request):
        """
        Identify PO users (PO Exam Center).
        PO == users in group "PO" or with role "PO_ADMIN".
        Superuser alone does NOT make a user PO, so OIC superusers still see CandidateProfile.
        """
        u = request.user
        in_po_group = u.groups.filter(name="PO").exists()
        has_po_role = getattr(u, "role", None) == "PO_ADMIN"
        return in_po_group or has_po_role


    def _field_exists(self, field_name: str) -> bool:
        """Check if a given field actually exists on CandidateProfile."""
        return any(f.name == field_name for f in CandidateProfile._meta.get_fields())

    def get_model_perms(self, request):
        """
        Hide CandidateProfile from the admin sidebar and app index for PO users,
        while still allowing them to hit the custom export URLs directly.
        """
        if self._is_po(request):
            return {}
        return super().get_model_perms(request)

    # ---------- changelist (top buttons/links area) ----------
    def changelist_view(self, request, extra_context=None):
        # Turn on inline editing only for PO (important: set attribute here)
        if self._is_po(request):
            self.list_editable = (
                "primary_viva_marks",
                "primary_practical_marks",
                # "secondary_viva_marks",
                # "secondary_practical_marks",
            )
        else:
            self.list_editable = ()  # no inline editing for others

        # IMPORTANT: do not inject export_all_* keys here (we remove top links client-side)
        return super().changelist_view(request, extra_context=extra_context)

    # ---------- list columns ----------
    def get_list_display(self, request):
        if self._is_po(request):
            # Minimal columns for PO with marks
            return (
                "army_no",
                "name",
                "trade",
                "primary_viva_marks",
                "primary_practical_marks",
                # "secondary_viva_marks",
                # "secondary_practical_marks",
            )
        # non-PO (your original default)
        return ("army_no", "name", "user", "rank", "trade", "shift", "created_at")

    # Remove row links for PO
    def get_list_display_links(self, request, list_display):
        if self._is_po(request):
            return ()  # no links → can't open detail page
        return super().get_list_display_links(request, list_display)

    # ---------- actions per role ----------
    def get_actions(self, request):
        actions = super().get_actions(request)
        if self._is_po(request):
            # PO can export DAT, Photos and Marks; hide other actions if you prefer
            return {
                k: v
                for k, v in actions.items()
                if k in ["export_candidates_dat", "export_candidate_images", "export_marks_excel"]
            }
        else:
            # Non-PO cannot export DAT, Photos or Marks via actions
            blocked = {"export_candidates_dat", "export_candidate_images", "export_marks_excel"}
            return {k: v for k, v in actions.items() if k not in blocked}

    # ---------- change form (add/change page) ----------
    def get_fields(self, request, obj=None):
        # Base fields for a full candidate (non-PO users)
        raw_base_fields = [
            "user",
            "army_no",
            "rank",
            "trade_type",
            "unit",
            "brigade",
            "corps",
            "command",
            "trade",
            "name",
            "dob",
            "doe",
            "aadhar_number",
            "father_name",
            "photograph",
            "med_cat",
            "cat",
            "nsqf_level",
            "exam_center",
            "training_center",
            "state",
            "district",
            "primary_qualification",
            "primary_duration",
            "primary_credits",
            # "secondary_qualification",
            # "secondary_duration",
            # "secondary_credits",
            "shift",
        ]
        base_fields = [f for f in raw_base_fields if self._field_exists(f)]

        po_only_fields_raw = [
            "primary_viva_marks",
            "primary_practical_marks",
            "secondary_viva_marks",
            "secondary_practical_marks",
        ]
        po_only_fields = [f for f in po_only_fields_raw if self._field_exists(f)]

        if self._is_po(request):
            # If someone hits the change URL directly, still show limited fields
            po_base = [f for f in ["army_no", "name", "trade"] if self._field_exists(f)]
            return po_base + po_only_fields

        # Non-PO: base fields + created_at if it exists
        if self._field_exists("created_at"):
            return base_fields + ["created_at"]
        return base_fields

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if self._is_po(request):
            # PO can edit only the four marks; everything else readonly
            all_possible = [
                "user",
                "army_no",
                "rank",
                "trade_type",
                "unit",
                "brigade",
                "corps",
                "command",
                "trade",
                "name",
                "dob",
                "doe",
                "aadhar_number",
                "father_name",
                "photograph",
                "med_cat",
                "cat",
                "nsqf_level",
                "exam_center",
                "training_center",
                "state",
                "district",
                "primary_qualification",
                "primary_duration",
                "primary_credits",
                # "secondary_qualification",
                # "secondary_duration",
                # "secondary_credits",
                "shift",
                "created_at",
            ]
            for f in all_possible:
                if self._field_exists(f) and f not in readonly:
                    readonly.append(f)
        else:
            if self._field_exists("created_at") and "created_at" not in readonly:
                readonly.append("created_at")
        return readonly

    # Block opening the change form UI for PO (must use list editing)
    def change_view(self, request, object_id, form_url="", extra_context=None):
        if self._is_po(request):
            return HttpResponseForbidden("PO edits marks on the list page only.")
        return super().change_view(request, object_id, form_url, extra_context)

    # prevent PO from add/delete
    def has_add_permission(self, request):
        if self._is_po(request):
            return False
        return super().has_add_permission(request)

    def has_delete_permission(self, request, obj=None):
        if self._is_po(request):
            return False
        return super().has_delete_permission(request, obj)

    # Optional per-object link (not used by PO because links are disabled)
    def download_csv_link(self, obj):
        url = reverse("export_candidate_pdf", args=[obj.id])
        return format_html('<a class="button" href="{}">Download Answers PDF</a>', url)

    download_csv_link.short_description = "Export PDF"
    download_csv_link.allow_tags = True

    # ---------- Custom URLs (only PO allowed) ----------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "Export-All-dat/",
                self.admin_site.admin_view(self.export_all_dat_view),
                name="registration_candidateprofile_export_all_dat",
            ),
            path(
                "Export-All-Images/",
                self.admin_site.admin_view(self.export_all_images_view),
                name="registration_candidateprofile_export_all_images",
            ),
            # NEW: Export-All-Marks (admin-bound method)
            path(
                "Export-All-Marks/",
                self.admin_site.admin_view(self.export_all_marks_view),
                name="registration_candidateprofile_export_all_marks",
            ),
            # JS endpoint that injects the sidebar buttons (served via admin view to allow permission check)
            # NOTE: we no longer serve the sidebar-injection JS; export is done from Dashboard.
        ]
        return custom_urls + urls

    def export_all_dat_view(self, request):
        # Only PO can export DAT
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        qs = self.get_queryset(request)
        return export_candidates_dat(self, request, qs)

    def export_all_images_view(self, request):
        # Only PO can export photos ZIP
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        return export_all_candidate_images(self, request)

    def export_all_marks_view(self, request):
        # Only PO can export Marks
        if not self._is_po(request):
            return HttpResponseForbidden("Not allowed.")
        qs = self.get_queryset(request)
        return export_marks_excel(self, request, qs)

    # No extra admin JS; export is accessed via Dashboard button only.